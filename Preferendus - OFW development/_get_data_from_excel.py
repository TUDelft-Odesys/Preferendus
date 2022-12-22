"""
File to get specific data from the Excel file
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook


class GetData:
    """
    class to get data from the master Excel file
    """

    def __init__(self):
        # read file
        self.wb = load_workbook('data_file.xlsx')

        vessel_tables = self._extract_data('Vessels')
        self.vessel_tasks_data = vessel_tables['vessel_data'].set_index('Vessel').to_dict('index')
        self.vessel_costs = vessel_tables['vessel_costs'].set_index('Vessel').to_dict('dict')['Day rate']
        self.vessel_costs_list = vessel_tables['vessel_costs']['Day rate'].values
        self.additional_costs = vessel_tables['additional_costs'].set_index('Additional costs').to_dict('dict')[
            'Day rate']
        self.vessel_fuel = vessel_tables['vessel_fuel_usage'].set_index('Vessel').to_dict('index')
        self.vessel_util = vessel_tables['vessel_util'].set_index('Vessel').values
        self.vessel_tables = vessel_tables

        project_tables = self._extract_data('Project_info')

        self.project_data = project_tables['project_info'].set_index('Identifier').to_dict()['Value']

        self.workability_limits = self._extract_data('Project_info')['workability_limits'].set_index('Task').to_dict()
        self.work_windows = self._extract_data('Project_info')['work_windows'].set_index('Task').to_dict()['Hours']

        self.soil_data = self._extract_data('Soil + Mooring data')['soil_data'].set_index('Item').to_dict()['value']
        self.mooring_data = self._extract_data('Soil + Mooring data')['mooring_data'].set_index('Item').to_dict()[
            'value']

    def _extract_data(self, operation: str):
        """

        :param operation:
        :return:
        """
        ws = self.wb[operation]
        mapping = {}
        for entry, data_boundary in ws.tables.items():
            # parse the data within the ref boundary
            data = ws[data_boundary]
            # extract the data
            # the inner list comprehension gets the values for each cell in the table
            content = [[cell.value for cell in ent]
                       for ent in data
                       ]

            header = content[0]

            # the contents ... excluding the header
            rest = content[1:]

            # create dataframe with the column names and pair table name with dataframe
            df = pd.DataFrame(rest, columns=header)
            mapping[entry] = df

        return mapping

    def get_governing_forces(self, water_depth):
        # initialize dictionary with results
        design_forces = dict()

        # extract data from master Excel
        forces_table = self._extract_data('Forces')['Forces']
        angle_seabed_table = self._extract_data('Angles')['AnglesML']
        relevant_data = forces_table[forces_table['Water depth'] == water_depth]
        relevant_data_angle_seabed = angle_seabed_table[angle_seabed_table['Water depth'] == water_depth]

        # divide data into the two mooring configurations
        data_taut = relevant_data[relevant_data['Mooring config'] == 'taut']
        data_catenary = relevant_data[relevant_data['Mooring config'] == 'catenary']

        # get data for catenary moorings
        forces_catenary_single = data_catenary[
            ['Max A1, single, factored', 'Max A2, single, factored', 'Max A3, single, factored']].values
        design_forces['catenary, single, governing'] = np.amax(forces_catenary_single)
        design_forces['catenary, angle'] = 0

        forces_catenary_shared = data_catenary[
            ['Max shared mooring, factored', 'Max shared mooring 2 anchors, factored']].values

        design_forces['catenary, shared, governing'] = np.amax(forces_catenary_shared)

        # get data for taut moorings
        forces_taut_single = data_taut[
            ['Max A1, single, factored', 'Max A2, single, factored', 'Max A3, single, factored']].values
        design_forces['taut, single, governing'] = np.amax(forces_taut_single)

        forces_taut_shared = data_taut[
            ['Max shared mooring, factored', 'Max shared mooring 2 anchors, factored']].values

        design_forces['taut, shared, governing'] = np.amax(forces_taut_shared)
        loc_shared = np.where(forces_taut_shared == np.amax(forces_taut_shared))[0]

        direction_design_force = data_taut['Direction'].values[loc_shared][0]
        directions_col = relevant_data_angle_seabed['Direction']
        loc_angle = np.where(directions_col == direction_design_force)[0]
        design_forces['taut, angle'] = relevant_data_angle_seabed['Angle ML wo force [rad]'][loc_angle].values[0]
        return design_forces

    def build_bounds_vessels(self):
        vessel_info = self.vessel_tables['vessel_data']
        masks = dict()
        names = list()

        mask_sp = vessel_info['SP installation'].values == 1
        masks['sp'] = mask_sp
        names += vessel_info['Vessel'][mask_sp].tolist()
        n_sp = sum(mask_sp)

        mask_pile = vessel_info['Pile installation'].values == 1
        masks['pile'] = mask_pile
        names += vessel_info['Vessel'][mask_pile].tolist()
        n_pile = sum(mask_pile)

        mask_dea = vessel_info['DEA installation'].values == 1
        masks['dea'] = mask_dea
        names += vessel_info['Vessel'][mask_dea].tolist()
        n_dea = sum(mask_dea)

        mask_ml_taut = vessel_info['ML poly'].values == 1
        masks['taut'] = mask_ml_taut
        names += vessel_info['Vessel'][mask_ml_taut].tolist()
        n_ml_taut = sum(mask_ml_taut)

        mask_ml_chain = vessel_info['ML chain'].values == 1
        masks['chain'] = mask_ml_chain
        names += vessel_info['Vessel'][mask_ml_chain].tolist()
        n_ml_chain = sum(mask_ml_chain)

        mask_ten = vessel_info['Tensioning'].values == 1
        masks['ten'] = mask_ten
        names += vessel_info['Vessel'][mask_ten].tolist()
        n_ten = sum(mask_ten)

        mask_hu = vessel_info['Hookup'].values == 1
        masks['hu'] = mask_hu
        names += vessel_info['Vessel'][mask_hu].tolist()
        n_hu = 1

        # mask_tow = vessel_info['Tow'].values == 1
        # names_tow = vessel_info['Vessel'][mask_tow].tolist()
        # n_tow = sum(mask_tow)

        bound_base = [[0, 1]]
        n_bounds_vessels = n_sp + n_pile + n_dea + n_ml_taut + n_ml_chain + n_ten
        bounds_vessels = bound_base * n_bounds_vessels
        bounds_vessels += [[0, int(np.count_nonzero(mask_hu) - 1)]]
        var_type = ['bool'] * (n_bounds_vessels + n_hu)

        return bounds_vessels, var_type, masks, names  # , n_tow, names_tow


if __name__ == '__main__':
    cls = GetData()

    cls.build_bounds_vessels()

    # pprint(cls.vessel_tasks_data)
    # pprint(cls.vessel_costs)
    # pprint(cls.vessel_fuel)
    #
    # pprint(cls.project_data)
