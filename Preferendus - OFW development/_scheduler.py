"""
Scheduler for floating wind case
"""
from calendar import month_abbr
from collections import Counter
from datetime import datetime, timedelta
from itertools import groupby
from pprint import pprint
from typing import Union

import numpy as np
import pandas as pd
import plotly.figure_factory as ff
from openpyxl import load_workbook
from scipy.optimize import brute

from schedule.drag_embedded_w_stevtensioning_anchors import DragEmbeddedAnchorsStevtensioningScheduler
from schedule.drag_embedded_wo_stevtensioning_anchors import DragEmbeddedAnchorsScheduler
from schedule.hookup import GetHookupSchedule
from schedule.intermediate_mobilization import GetInterMobSchedule
from schedule.mooring_line_install import GetMooringLegInstallSchedule
from schedule.pilled_anchors import PilledAnchorsScheduler
from schedule.suction_anchors import SuctionAnchorsScheduler
from schedule.tensioning import GetTensioningSchedule


class CreateSchedule:
    """
    class to create schedule

    https://bit.ly/3s7rR5r
    """

    def __init__(self, hookup_version='full'):
        # read file
        self.wb = load_workbook('data_file.xlsx')
        self.schedule = list()

        project_data_sheet = self._extract_data('Project_info')
        self.project_info = project_data_sheet['project_info'].set_index('Identifier').to_dict()['Value']
        vessel_data_sheet = self._extract_data('Vessels')
        self.vessel_info = vessel_data_sheet['vessel_data']

        self.times = self._get_times()

        anchor_schedules = dict()
        anchor_schedules['dea_stev'] = DragEmbeddedAnchorsStevtensioningScheduler(
            self._extract_data('DEA w stevtensioning'), self.times)
        anchor_schedules['dea_non-stev'] = DragEmbeddedAnchorsScheduler(
            self._extract_data('DEA wo stevtensioning'), self.times)
        anchor_schedules['pilled'] = PilledAnchorsScheduler(self._extract_data('Anchor Piles'), self.times)
        anchor_schedules['suction'] = SuctionAnchorsScheduler(self._extract_data('Suction Anchors'), self.times)

        self.anchor_schedules = anchor_schedules
        self.inter_mobs = GetInterMobSchedule(self._extract_data('General'), self.times)
        self.mooring_line_schedule = GetMooringLegInstallSchedule(self._extract_data('Mooring Leg Install'), self.times)
        self.tensioning_schedule = GetTensioningSchedule(self._extract_data('Tensioning'), self.times)
        self.hookup_schedule = GetHookupSchedule(self._extract_data('Hookup'), self._extract_data('Towing'),
                                                 hookup_version, self.times)

        self.wait_times = self._wait_time_ml_anchor()
        self.s_data_basis = {
            'idle': 0,
            'port': 0,
            'transit': 0,
            'DP': 0,
            'AH': 0,
            'towing': 0
        }

    def _get_times(self):
        """
        Local distances: https://www.bednblue.com/sailing-distance-calculator
        long distances: http://ports.com/sea-route/

        :return:
        """
        speeds = {  # in knots
            'transit': 13,
            'DP /wo ROV': 2,
            'DP /w ROV': 0.5,
            'towing': 3
        }

        distances = {
            'home - marshalling yard': self.project_info['Distance home - marshalling yard'],  # R'dam - Ulsan
            'home - fabrication yard': self.project_info['Distance home - fabrication yard'],  # R'dam - Busan
            'marshalling yard - site': self.project_info['Distance marshalling yard - site'],  # Ulsan - site
            'fabrication yard - site': self.project_info['Distance fabrication yard - site'],  # Busan - site
        }

        times = {
            'transit home - marshalling yard': distances['home - marshalling yard'] / speeds['transit'] / 24,
            'transit home - fabrication yard': distances['home - fabrication yard'] / speeds['transit'] / 24,
            'transit to site': distances['marshalling yard - site'] / speeds['transit'] / 24,
            'towing': distances['fabrication yard - site'] / speeds['towing'] / 24,
            'transit site - fabrication yard': distances['fabrication yard - site'] / speeds['transit'] / 24,
            'in-site': 10 / 60,
            'loading time per ML': 5 / 60,
            'loading time per anchor: dea_stev': 10 / 60,
            'loading time per anchor: dea_non-stev': 10 / 60,
            'loading time per anchor: pilled': 5 / 60,
            'loading time per anchor: suction': 5 / 60,
            'loading time per anchor: drilled': 5 / 60,
        }

        return times

    def _extract_data(self, sheet: str):
        """
        :param sheet: name of worksheet
        :return:
        """
        ws = self.wb[sheet]
        mapping = {}
        for entry, data_boundary in ws.tables.items():
            # parse the data within the ref boundary
            data = ws[data_boundary]
            # extract the data
            # the inner list comprehension gets the values for each cell in the table
            content = [[cell.value for cell in ent]
                       for ent in data
                       ]

            header = content[0]

            # the contents excluding the header
            rest = content[1:]

            # create dataframe with the column names and pair table name with dataframe
            df = pd.DataFrame(rest, columns=header)
            mapping[entry] = df

        return mapping

    def _get_monthly_workability(self, workability_criteria: dict, work_windows: dict):
        """
        Extract workability figures for certain criteria.

        :param workability_criteria: dict with workability limits per env condition
        :param work_windows: dict of hours that it should be lower than the given limit
        :return: None
        """
        project = self.project_info['Name']

        df_wind = pd.DataFrame({'Workability': workability_criteria['wind']})
        sequential_per_month_wind = dict()
        for month in np.arange(1, 13):
            workability_per_month = list()
            csv_import = pd.read_csv(f'./env_analysis/figures/{project}/Wind/csv/Wind_data_month_{month}.csv')
            data = csv_import[month_abbr[month]].tolist()
            sequential_per_limit = dict()

            for limit in workability_criteria['wind']:
                mask = np.array(data) < limit
                counter = np.count_nonzero(mask)
                percentage = counter / len(data) * 100
                workability_per_month.append(round(percentage, 2))

                counts = Counter(sum(1 for _ in g) for k, g in groupby(mask) if k == 1)
                sequential_per_seq = dict()
                for seq in work_windows:
                    valid_keys = [key for key in counts if key >= seq]
                    count = 0
                    for item in valid_keys:
                        count += item * counts[item]
                    sequential_per_seq[seq] = round(count / len(data) * 100, 2)

                sequential_per_limit[limit] = sequential_per_seq

            df_wind[month_abbr[month]] = workability_per_month
            sequential_per_month_wind[month] = sequential_per_limit

        df_wind.to_csv(f'./env_analysis/figures/{project}/Wind/csv/Wind_workability.csv')

        df_waves = pd.DataFrame({'Workability': workability_criteria['waves']})
        sequential_per_month_waves = dict()
        for month in np.arange(1, 13):
            workability_per_month = list()
            csv_import = pd.read_csv(f'./env_analysis/figures/{project}/Waves/csv/Wave_data_month_{month}.csv')
            data = csv_import[month_abbr[month]].tolist()
            sequential_per_limit = dict()

            for limit in workability_criteria['waves']:
                mask = np.array(data) < limit
                counter = np.count_nonzero(mask)
                percentage = counter / len(data) * 100
                workability_per_month.append(round(percentage, 2))

                counts = Counter(sum(1 for _ in g) for k, g in groupby(mask) if k == 1)
                sequential_per_seq = dict()
                for seq in work_windows:
                    valid_keys = [key for key in counts if key >= seq]
                    count = 0
                    for item in valid_keys:
                        count += item * counts[item]
                    sequential_per_seq[seq] = round(count / len(data) * 100, 2)

                sequential_per_limit[limit] = sequential_per_seq

            df_waves[month_abbr[month]] = workability_per_month
            sequential_per_month_waves[month] = sequential_per_limit

        df_waves.to_csv(f'./env_analysis/figures/{project}/Waves/csv/Waves_workability.csv')

        work_windows_wind = dict()
        for limit in workability_criteria['wind']:
            work_windows_wind[limit] = dict()
            for key in sequential_per_month_wind:
                work_windows_wind[limit][key] = sequential_per_month_wind[key][limit]

        work_windows_waves = dict()
        for limit in workability_criteria['waves']:
            work_windows_waves[limit] = dict()
            for key in sequential_per_month_waves:
                work_windows_waves[limit][key] = sequential_per_month_waves[key][limit]

        return work_windows_wind, work_windows_waves

    def workability_analysis(self, workability_percentage, workability_criteria: dict, work_windows):
        """
        Returns workable months for a given set of workability criteria and work windows
        """

        work_windows_wind, work_windows_waves = self._get_monthly_workability(workability_criteria, work_windows)

        workable_months_per_limit_wind = list()
        workable_months_per_limit_waves = list()

        for criteria in workability_criteria['wind']:
            data = work_windows_wind[criteria]
            workable_months = dict()
            for window in work_windows:
                workable_months[window] = list()
                for month in np.arange(1, 13):
                    if data[month][window] >= workability_percentage:
                        workable_months[window].append(month)
            workable_months_per_limit_wind.append(workable_months)

        for criteria in workability_criteria['waves']:
            data = work_windows_waves[criteria]
            workable_months = dict()
            for window in work_windows:
                workable_months[window] = list()
                for month in np.arange(1, 13):
                    if data[month][window] >= workability_percentage:
                        workable_months[window].append(month)
            workable_months_per_limit_waves.append(workable_months)

        return workable_months_per_limit_wind, workable_months_per_limit_waves

    def _wait_time_ml_anchor(self):
        """
        The times a vessel has to wait on another in simultaneous operations
        :return:
        """
        wait_times = dict()

        duration_pile_install = sum(
            self.anchor_schedules['pilled'].building_blocks['anchor_installation']['time'][0:11])
        duration_sa_install = sum(self.anchor_schedules['suction'].building_blocks['anchor_installation']['time'][0:11])
        duration_ml_chain = sum(self.mooring_line_schedule.building_blocks['ml_install_chain']['time'][0:9])
        duration_ml_poly = sum(self.mooring_line_schedule.building_blocks['ml_install_poly']['time'][0:15])

        wait_times['ml-pile, chain'] = duration_pile_install - duration_ml_chain
        wait_times['ml-sa, chain'] = duration_sa_install - duration_ml_chain
        wait_times['ml-pile, poly'] = duration_pile_install - duration_ml_poly
        wait_times['ml-sa, poly'] = duration_sa_install - duration_ml_poly

        return wait_times

    def _sus_time(self, data_array):
        """
        Get fractions of the time a vessel is performing an activiy
        """
        s_times = list()
        total_time = sum(data_array['time'])
        for key in self.s_data_basis.keys():
            task_time = sum(data_array[data_array['emission-cat'] == key]['time'])
            s_times.append(task_time / total_time)

        return np.array(s_times)

    def _anchor_install(self, anchor_type, ml_type, inter_mob_variant, vessels, deck_length, length, diameter,
                        anchors_to_install, start_season, end_season, max_ml_chain_onboard):
        """
        Get times etc. of anchor installation
        """

        if 'dea' in anchor_type:
            space_a = min(length, diameter)
        elif 'suction' in anchor_type:  # SPs stand vertical onboard
            space_a = diameter
        else:  # anchor piles lie horizontal onboard
            space_a = length

        if 'dea' in anchor_type:
            max_anchors_onboard = np.amin([max_ml_chain_onboard, 2 * (deck_length // space_a)], axis=0)
        else:
            max_anchors_onboard = 2 * (deck_length // space_a)

        if sum(np.multiply(max_anchors_onboard, vessels)) == 0.:
            max_anchors_onboard = [1] * len(vessels)

        # print('vessels:', vessels)
        # print('max a onboard:', max_anchors_onboard)

        # determine anchors to be installed per iteration
        anchors_installed_per_iter = np.sum(np.multiply(vessels, max_anchors_onboard), axis=0, dtype=float)
        if anchors_installed_per_iter == 0:
            anchors_installed_per_iter = 1e-6  # prevent div by 0 problems

        # determine number of iterations needed
        iterations = anchors_to_install / anchors_installed_per_iter

        # determine the number of days in and in between an installation season
        days_in_season = (end_season - start_season).days
        days_between_seasons = 365 - days_in_season

        # check if integrated is in name and take it out
        if '-integrated' in anchor_type:
            anchor_type_base = anchor_type.replace('-integrated', '')
        else:
            anchor_type_base = anchor_type

        # get data of process
        data_anchor = self.anchor_schedules[anchor_type_base].building_blocks['anchor_installation'].copy()

        # determine idle time of ship when integrated installation of anchor and ML is done
        idle_time = 0
        if '-integrated' in anchor_type:
            ml = 'poly' if ml_type == 'taut' else 'chain'
            anchor = 'sa' if anchor_type == 'suction-integrated' else 'pile'
            if self.wait_times[f'ml-{anchor}, {ml}'] < 0:  # anchor install vessel has to wait on ML install vessel
                idle_time = -1 * self.wait_times[f'ml-{anchor}, {ml}']

            data_anchor['time'][4] = idle_time

        # get total installation time
        install_time_anchors = sum(data_anchor['time'])
        sus_times_install = self._sus_time(data_anchor)

        # get time for the type of intermediate mobilization and mobilization (depending on ML on board)
        mobilization = list()
        intermediate_mobilization = list()
        sus_times_mob_per_vessel = list()
        sus_times_inter_mob_per_vessel = list()
        for max_a in max_anchors_onboard:
            mob_bare = self.anchor_schedules[anchor_type_base].building_blocks['mobilization'].copy()
            if inter_mob_variant:
                inter_mob_bare = self.inter_mobs.building_blocks['offshore_bunkering'].copy()
            else:
                inter_mob_bare = self.inter_mobs.building_blocks['intermediate_mobilization'].copy()

            mob_bare['time'][np.where(mob_bare == 'var')[0][0]] = max_a * self.times[
                f'loading time per anchor: {anchor_type_base}']
            inter_mob_bare['time'][np.where(inter_mob_bare == 'var')[0][0]] = max_a * self.times[
                f'loading time per anchor: {anchor_type_base}']

            mobilization.append(sum(mob_bare['time']))
            intermediate_mobilization.append(sum(inter_mob_bare['time']))
            sus_times_mob_per_vessel.append(self._sus_time(mob_bare))
            sus_times_inter_mob_per_vessel.append(self._sus_time(inter_mob_bare))

        # determine run time per iteration
        run_time_per_iteration = np.multiply(max_anchors_onboard, install_time_anchors)

        def objective(x):
            """Objective to determine the number of iterations per vessel is the most efficient given the capacity"""
            if np.sum(np.multiply(np.round_(x), max_anchors_onboard)) < anchors_to_install:
                return 1e10
            else:
                return np.sum(np.multiply(x, max_anchors_onboard))

        ranges = ()
        for item in vessels:
            if item > 0:
                ranges += (slice(0, np.ceil(iterations) + 1, 1),)
            else:
                ranges += (slice(0, 1, 1),)
        iterations_per_vessel_type = brute(objective, ranges, finish=None)
        overshoot = np.sum(np.multiply(iterations_per_vessel_type, max_anchors_onboard)) - anchors_to_install

        number_of_mob_per_vessel = np.array(iterations_per_vessel_type) - 1
        mask = number_of_mob_per_vessel < 0
        number_of_mob_per_vessel[mask] = 0

        # print('iter:', iterations_per_vessel_type)
        # print('install time', install_time_anchors)

        number_of_inter_mobs = np.multiply(number_of_mob_per_vessel, vessels)
        number_of_vessel_movement = np.multiply(iterations_per_vessel_type, vessels)

        project_duration = np.multiply(run_time_per_iteration, iterations_per_vessel_type) + \
                           np.multiply(number_of_inter_mobs, intermediate_mobilization)

        run_time_install = np.multiply(run_time_per_iteration, number_of_vessel_movement)
        run_time_inter_mob = np.multiply(number_of_inter_mobs, intermediate_mobilization)
        time_per_ship = np.sum([run_time_install, run_time_inter_mob], axis=0)

        anchors_installed_per_vessel = np.multiply(max_anchors_onboard, iterations_per_vessel_type)

        if overshoot > 0:
            overshoot_per_vessel = np.round_(anchors_installed_per_vessel / sum(anchors_installed_per_vessel) * overshoot)
            # print(f'overshoot per vessel: {overshoot_per_vessel}')
            mask1 = anchors_installed_per_vessel > 0
            try:
                if sum(overshoot_per_vessel) != overshoot and sum(overshoot_per_vessel) % overshoot == sum(
                        overshoot_per_vessel):
                    overshoot_per_vessel[mask1][0] += abs(sum(overshoot_per_vessel) - overshoot)
                elif sum(overshoot_per_vessel) != overshoot:
                    overshoot_per_vessel[mask1][0] -= abs(sum(overshoot_per_vessel) - overshoot)
            except IndexError as err:
                print(anchors_installed_per_vessel)
                print(overshoot_per_vessel)
                print(sum(overshoot_per_vessel), overshoot)
                raise err
            too_much_installed = overshoot_per_vessel * install_time_anchors
            project_duration -= too_much_installed
            time_per_ship -= (too_much_installed * vessels)
            anchors_installed_per_vessel -= overshoot_per_vessel

        # print(f'anchors_installed_per_vessel: {anchors_installed_per_vessel}')
        """
        Note time per ship is summed per number of vessels, whereas the project duration is considering the vessels 
        working in parallel. Hence, the time per vessel can be higher than the project duration!
        """

        max_mob = max(np.multiply(mobilization, vessels))

        demob_time = sum(self.anchor_schedules[anchor_type_base].building_blocks['demobilization']['time'])
        sus_demob = self._sus_time(self.anchor_schedules[anchor_type_base].building_blocks['demobilization'])

        # print('time per ship:', time_per_ship)
        # print('mob:', mobilization)
        # print('demob:', demob_time)

        extra_seasons = np.floor(project_duration / days_in_season)
        # print('extra seasons:', extra_seasons)
        run_time_ships = time_per_ship + (extra_seasons + 1) * (
                np.array(mobilization) + demob_time) * vessels

        # compensate for first idle time
        mask = run_time_ships != 0
        run_time_ships[mask] -= idle_time

        project_duration += extra_seasons * days_between_seasons + (extra_seasons + 1) * (
                max_mob + demob_time)

        sus_time_per_ship = list()
        for it, time in enumerate(time_per_ship):
            frac_install = np.multiply(sus_times_install, run_time_install[it])
            frac_inter_mob = np.multiply(sus_times_inter_mob_per_vessel[it], run_time_inter_mob[it])
            frac_mob = np.multiply(mobilization[it] * vessels[it] * (extra_seasons[it] + 1),
                                   sus_times_mob_per_vessel[it])
            frac_demob = np.multiply(demob_time * vessels[it] * (extra_seasons[it] + 1), sus_demob)
            sus_time_per_ship.append(np.sum([frac_install, frac_inter_mob, frac_mob, frac_demob], axis=0))

        s_data = self.s_data_basis.copy()
        for it, key in enumerate(s_data.keys()):
            s_data[key] = np.array(sus_time_per_ship)[:, it]

        # print('proj duration:', max(project_duration))
        return max(project_duration), run_time_ships, extra_seasons, max_mob, demob_time, number_of_inter_mobs, \
               intermediate_mobilization, s_data, anchors_installed_per_vessel

    def _mooring_line_install(self, anchor_type, ml_type, vessels, max_ml_onboard, ml_to_install,
                              start_season, end_season, anchors_installed_per_vessel):
        """
        Get times etc. of anchor installation
        """

        # determine ml to be installed per iteration
        ml_installed_per_iter = np.sum(np.multiply(vessels, max_ml_onboard), axis=0, dtype=float)
        if ml_installed_per_iter == 0:
            ml_installed_per_iter = 1e-6  # prevent div by 0 problems

        # print('vessels:', vessels)
        # print('max ml onboard:', max_ml_onboard)

        # determine number of iterations needed
        iterations = ml_to_install / ml_installed_per_iter

        # determine the number of days in and in between an installation season
        days_in_season = (end_season - start_season).days
        days_between_seasons = 365 - days_in_season

        # get data of process
        if ml_type == 'taut':
            data_ml = self.mooring_line_schedule.building_blocks['ml_install_poly'].copy()
        else:
            data_ml = self.mooring_line_schedule.building_blocks['ml_install_chain'].copy()

        # determine idle time of ship when integrated installation of anchor and ML is done
        idle_time = 0
        if '-integrated' in anchor_type:
            ml = 'poly' if ml_type == 'taut' else 'chain'
            anchor = 'sa' if anchor_type == 'suction-integrated' else 'pile'
            if self.wait_times[f'ml-{anchor}, {ml}'] > 0:  # ML ship has to wait on anchor install
                idle_time = self.wait_times[f'ml-{anchor}, {ml}']

            data_ml['time'][4] = idle_time

        # get total installation time
        install_time_ml = sum(data_ml['time'])
        sus_times_install = self._sus_time(data_ml)

        # get time for the type of intermediate mobilization and mobilization (depending on ML on board)
        mobilization = list()
        intermediate_mobilization = list()
        sus_times_mob_per_vessel = list()
        sus_times_inter_mob_per_vessel = list()
        for max_ml in max_ml_onboard:
            mob_bare = self.mooring_line_schedule.building_blocks['mobilization'].copy()
            inter_mob_bare = self.inter_mobs.building_blocks['intermediate_mobilization'].copy()

            mob_bare['time'][2] = max_ml * self.times['loading time per ML']
            inter_mob_bare['time'][1] = max_ml * self.times['loading time per ML']

            mobilization.append(sum(mob_bare['time']))
            intermediate_mobilization.append(sum(inter_mob_bare['time']))
            sus_times_mob_per_vessel.append(self._sus_time(mob_bare))
            sus_times_inter_mob_per_vessel.append(self._sus_time(inter_mob_bare))

        # determine run time per iteration
        run_time_per_iteration = np.multiply(max_ml_onboard, install_time_ml)
        ml_onboard_total = np.multiply(vessels, max_ml_onboard)  # total number of anchors on board all vessels

        def objective(x):
            """Objective to determine the number of iterations per vessel is the most efficient given the capacity"""
            if np.sum(np.multiply(np.round_(x), ml_onboard_total)) < ml_to_install:
                return 1e8
            else:
                return np.sum(np.multiply(x, ml_onboard_total))

        if '-integrated' in anchor_type:
            if np.count_nonzero(vessels) != np.count_nonzero(anchors_installed_per_vessel) or sum(
                    anchors_installed_per_vessel) == 0:
                iterations_per_vessel_type = np.array([1000.] * len(vessels))
            else:
                iterations_per_vessel_type = np.zeros_like(vessels)
                anchors_to_ml = anchors_installed_per_vessel / sum(anchors_installed_per_vessel) * self.project_info[
                    'Anchors to install']

                anchors_to_ml.sort()
                ml_onboard = np.multiply(vessels, max_ml_onboard)
                sort_index = np.argsort(ml_onboard)

                # print(anchors_to_ml)
                # print(ml_onboard)

                for index, j in enumerate(sort_index[::-1], start=1):
                    if ml_onboard[j] != 0:
                        iterations_per_vessel_type[j] = anchors_to_ml[-1 * index] / ml_onboard[j]
                iterations_per_vessel_type *= vessels
        else:
            ranges = ()
            for item in vessels:
                if item > 0:
                    ranges += (slice(0, np.ceil(iterations) + 1, 1),)
                else:
                    ranges += (slice(0, 1, 1),)
            iterations_per_vessel_type = brute(objective, ranges, finish=None)

        # print('iter', iterations_per_vessel_type)
        # print('install time', install_time_ml, intermediate_mobilization)
        overshoot = np.sum(np.multiply(iterations_per_vessel_type, max_ml_onboard)) - ml_to_install

        number_of_mob_per_vessel = np.array(iterations_per_vessel_type) - 1
        mask = number_of_mob_per_vessel < 0
        number_of_mob_per_vessel[mask] = 0

        number_of_inter_mobs = np.multiply(number_of_mob_per_vessel, vessels)
        number_of_vessel_movement = np.multiply(iterations_per_vessel_type, vessels)

        project_duration = np.multiply(run_time_per_iteration, iterations_per_vessel_type) + \
                           np.multiply(number_of_inter_mobs, intermediate_mobilization)

        run_time_install = np.multiply(run_time_per_iteration, number_of_vessel_movement)
        run_time_inter_mob = np.multiply(number_of_inter_mobs, intermediate_mobilization)
        time_per_ship = np.sum([run_time_install, run_time_inter_mob], axis=0)

        ml_installed_per_vessel = np.multiply(max_ml_onboard, iterations_per_vessel_type)
        # print(f'ml_installed_per_vessel: {ml_installed_per_vessel}')
        if overshoot > 0:
            overshoot_per_vessel = np.round_(
                ml_installed_per_vessel / sum(ml_installed_per_vessel) * overshoot)
            mask1 = ml_installed_per_vessel > 0
            if sum(overshoot_per_vessel) != overshoot and sum(overshoot_per_vessel) % overshoot == sum(
                    overshoot_per_vessel):
                overshoot_per_vessel[mask1][0] += abs(sum(overshoot_per_vessel) - overshoot)
            elif sum(overshoot_per_vessel) != overshoot:
                overshoot_per_vessel[mask1][0] -= abs(sum(overshoot_per_vessel) - overshoot)
            # print(f'overshoot_per_vessel: {overshoot_per_vessel}')
            too_much_installed = overshoot_per_vessel * install_time_ml
            project_duration -= too_much_installed
            time_per_ship -= (too_much_installed * vessels)
            try:
                ml_installed_per_vessel -= overshoot_per_vessel
            except:
                print(ml_installed_per_vessel, type(ml_installed_per_vessel))
                print(overshoot_per_vessel, type(overshoot_per_vessel))
                raise BaseException
        # print(f'ml_installed_per_vessel: {ml_installed_per_vessel}')

        """
        Note time per ship is summed per number of vessels, whereas the project duration is considering the vessels 
        working in parallel. Hence, the time per vessel can be higher than the project duration!
        """

        # print('ship time', time_per_ship)
        max_mob = max(np.multiply(mobilization, vessels))

        demob_time = sum(self.mooring_line_schedule.building_blocks['demobilization']['time'])
        sus_demob = self._sus_time(self.mooring_line_schedule.building_blocks['demobilization'])
        # print('mob', mobilization)
        # print('demob', demob_time)

        extra_seasons = np.floor(project_duration / days_in_season)
        # print('extra seasons', extra_seasons)
        run_time_ships = time_per_ship + (extra_seasons + 1) * (
                np.array(mobilization) + demob_time) * vessels

        # compensate for first idle time
        mask = run_time_ships != 0
        run_time_ships[mask] -= idle_time

        project_duration += extra_seasons * days_between_seasons + (extra_seasons + 1) * (
                max_mob + demob_time)

        sus_time_per_ship = list()
        for it, time in enumerate(time_per_ship):
            frac_install = np.multiply(sus_times_install, run_time_install[it])
            frac_inter_mob = np.multiply(sus_times_inter_mob_per_vessel[it], run_time_inter_mob[it])
            frac_mob = np.multiply(mobilization[it] * vessels[it] * (extra_seasons[it] + 1),
                                   sus_times_mob_per_vessel[it])
            frac_demob = np.multiply(demob_time * vessels[it] * (extra_seasons[it] + 1), sus_demob)
            sus_time_per_ship.append(np.sum([frac_install, frac_inter_mob, frac_mob, frac_demob], axis=0))

        s_data = self.s_data_basis.copy()
        for it, key in enumerate(s_data.keys()):
            s_data[key] = np.array(sus_time_per_ship)[:, it]

        # print(max(project_duration))
        return max(project_duration), run_time_ships, extra_seasons, max_mob, demob_time, number_of_inter_mobs, \
               intermediate_mobilization, s_data

    def _tensioning(self, vessels, ml_to_tension, start_season, end_season):

        # determine number of iterations needed
        if np.sum(vessels, axis=0, dtype=float) == 0:
            iterations = 1e3
        else:
            iterations = ml_to_tension / np.sum(vessels, axis=0, dtype=float)

        # print('tensioning')
        # print('vessels', vessels)
        # determine the number of days in and in between an installation season
        days_in_season = (end_season - start_season).days
        days_between_seasons = 365 - days_in_season

        # get data of process
        data_tensioning = self.tensioning_schedule.building_blocks['tensioning']

        # get total installation time
        tensioning_time = sum(data_tensioning['time'])
        sus_times_install = self._sus_time(data_tensioning)

        project_duration = np.multiply(tensioning_time, iterations)

        iterations_per_vessel_type = np.multiply(vessels, iterations)
        # print('iter', iterations_per_vessel_type)
        # print('ten time', tensioning_time)
        time_per_ship = np.multiply(tensioning_time, iterations_per_vessel_type)
        # print('ship_time', time_per_ship)

        """
        Note time per ship is summed per number of vessels, whereas the project duration is considering the vessels 
        working in parallel. Hence, the time per vessel can be higher than the project duration!
        """

        mob_time = sum(self.tensioning_schedule.building_blocks['mobilization']['time'])
        sus_mob = self._sus_time(self.tensioning_schedule.building_blocks['mobilization'])
        # print('mob', mob_time)

        demob_time = sum(self.tensioning_schedule.building_blocks['demobilization']['time'])
        sus_demob = self._sus_time(self.tensioning_schedule.building_blocks['demobilization'])
        # print('demob', demob_time)

        extra_seasons = np.floor(project_duration / days_in_season)
        # print('extra seasons', extra_seasons)
        run_time_ships = time_per_ship + (extra_seasons + 1) * (mob_time + demob_time) * vessels
        project_duration += extra_seasons * days_between_seasons + (extra_seasons + 1) * (mob_time + demob_time)

        sus_time_per_ship = list()
        for it, time in enumerate(time_per_ship):
            frac_install = np.multiply(sus_times_install, time_per_ship[it])
            frac_mob = np.multiply(mob_time * vessels[it] * (extra_seasons + 1), sus_mob)
            frac_demob = np.multiply(demob_time * vessels[it] * (extra_seasons + 1), sus_demob)
            sus_time_per_ship.append(np.sum([frac_install, frac_mob, frac_demob], axis=0))

        s_data = self.s_data_basis.copy()
        for it, key in enumerate(s_data.keys()):
            s_data[key] = np.array(sus_time_per_ship)[:, it]

        # print('proj duration', project_duration)
        return project_duration, run_time_ships, extra_seasons, mob_time, demob_time, s_data

    def _hookup_towing(self, hu_vessels, turbines_to_install, start_season, end_season):
        # determine number of iterations needed
        n_vessels = 1  # np.sum(hu_vessels, axis=0, dtype=float)
        if n_vessels == 0.:
            n_vessels = 1e-6
        iterations = turbines_to_install / n_vessels

        # print('iter', iterations)

        # determine the number of days in and in between an installation season
        days_in_season = (end_season - start_season).days
        days_between_seasons = 365 - days_in_season

        # get data of process
        data_hookup = self.hookup_schedule.building_blocks['hookup']
        data_towing = self.hookup_schedule.building_blocks['towing']

        hookup_time_ex_idle = sum(data_hookup[data_hookup.task != 'Waiting on FWT arrival']['time'])
        tow_time_ex_idle = sum(data_towing[data_towing.task != 'Idle']['time'])
        max_process_time = max(hookup_time_ex_idle, tow_time_ex_idle)

        idle_ideal = abs(hookup_time_ex_idle - tow_time_ex_idle)
        if hookup_time_ex_idle > tow_time_ex_idle:
            wait_time_tow = idle_ideal
            wait_time_hu = 0
        else:
            wait_time_tow = 0
            wait_time_hu = idle_ideal

        wait_time_turbine_delivery = n_vessels * self.project_info['Days between complete FWT'] - max_process_time
        if wait_time_turbine_delivery < 0:
            wait_time_turbine_delivery = 0

        wait_time_hu += wait_time_turbine_delivery
        wait_time_tow += wait_time_turbine_delivery

        data_hookup['time'][1] = wait_time_hu
        data_towing['time'][5] = wait_time_tow

        # get total installation time
        hookup_time = sum(data_hookup['time'])
        towing_time = sum(data_towing['time'])
        sus_times_hu = self._sus_time(data_hookup)
        sus_times_tow = self._sus_time(data_towing)

        project_duration = np.multiply(hookup_time, iterations)
        iterations_per_vessel_type = np.multiply([1], iterations)

        time_per_hu_ship = np.multiply(hookup_time, iterations_per_vessel_type)
        # print('ship times', time_per_hu_ship)
        time_per_tow_ship = np.multiply(towing_time, 2 * iterations_per_vessel_type)

        """
        Note time per ship is summed per number of vessels, whereas the project duration is considering the vessels 
        working in parallel. Hence, the time per vessel can be higher than the project duration!
        """

        mob_time_hu = sum(self.hookup_schedule.building_blocks['mobilization HUV']['time'])
        sus_mob_hu = self._sus_time(self.hookup_schedule.building_blocks['mobilization HUV'])
        demob_time_hu = sum(self.hookup_schedule.building_blocks['demobilization HUV']['time'])
        sus_demob_hu = self._sus_time(self.hookup_schedule.building_blocks['demobilization HUV'])

        # print('mob', mob_time_hu)
        # print('demob', demob_time_hu)

        mob_time_tow = sum(self.hookup_schedule.building_blocks['mobilization Tug']['time'])
        sus_mob_tow = self._sus_time(self.hookup_schedule.building_blocks['mobilization Tug'])
        demob_time_tow = sum(self.hookup_schedule.building_blocks['demobilization Tug']['time'])
        sus_demob_tow = self._sus_time(self.hookup_schedule.building_blocks['demobilization Tug'])

        extra_seasons = np.floor(project_duration / days_in_season)
        # print('days in season', days_in_season)
        # print('HU time', hookup_time)

        run_time_hu_ships = time_per_hu_ship + (extra_seasons + 1) * (mob_time_hu + demob_time_hu) * np.array([1])

        # compensate for first idle time, however, still account for delivery rate FWTs
        mask = run_time_hu_ships != 0
        run_time_hu_ships[mask] -= wait_time_hu

        run_time_tow_ships = time_per_tow_ship + (extra_seasons + 1) * (mob_time_tow + demob_time_tow) * 2 * hu_vessels

        max_mob = max(mob_time_tow, mob_time_hu)
        max_demob = max(demob_time_tow, demob_time_hu)

        project_duration += extra_seasons * days_between_seasons + (extra_seasons + 1) * (
                max_mob + max_demob)

        sus_time_per_hu_ship = list()
        for it, time in enumerate(time_per_hu_ship):
            frac_install = np.multiply(sus_times_hu, time_per_hu_ship[it])
            frac_mob = np.multiply(mob_time_hu * hu_vessels[it] * (extra_seasons + 1), sus_mob_hu)
            frac_demob = np.multiply(demob_time_hu * hu_vessels[it] * (extra_seasons + 1), sus_demob_hu)
            sus_time_per_hu_ship.append(np.sum([frac_install, frac_mob, frac_demob], axis=0))

        s_data_hu = self.s_data_basis.copy()
        for it, key in enumerate(s_data_hu.keys()):
            s_data_hu[key] = np.array(sus_time_per_hu_ship)[:, it]

        sus_time_per_tow_ship = list()
        for it, time in enumerate(time_per_tow_ship):
            frac_install = np.multiply(sus_times_tow, time_per_tow_ship[it])
            frac_mob = np.multiply(mob_time_tow * hu_vessels[it] * 2 * (extra_seasons + 1), sus_mob_tow)
            frac_demob = np.multiply(demob_time_tow * hu_vessels[it] * 2 * (extra_seasons + 1), sus_demob_tow)
            sus_time_per_tow_ship.append(np.sum([frac_install, frac_mob, frac_demob], axis=0))

        s_data_tow = self.s_data_basis.copy()
        for it, key in enumerate(s_data_tow.keys()):
            s_data_tow[key] = np.array(sus_time_per_tow_ship)[:, it]

        # print(project_duration)
        return project_duration, run_time_hu_ships, run_time_tow_ships, extra_seasons, max_mob, max_demob, s_data_hu, \
               s_data_tow

    def _check_pop_size(self):
        summation = 0

        mask_sp = self.vessel_info['SP installation'] == 1
        summation += np.count_nonzero(mask_sp)

        mask_pile = self.vessel_info['Pile installation'] == 1
        summation += np.count_nonzero(mask_pile)

        mask_dea = self.vessel_info['DEA installation'] == 1
        summation += np.count_nonzero(mask_dea)

        mask_ml_taut = self.vessel_info['ML poly'] == 1
        summation += np.count_nonzero(mask_ml_taut)

        mask_ml_chain = self.vessel_info['ML chain'] == 1
        summation += np.count_nonzero(mask_ml_chain)

        mask_ten = self.vessel_info['Tensioning'] == 1
        summation += np.count_nonzero(mask_ten)

        mask_hu = self.vessel_info['Hookup'] == 1
        summation += 1

        # mask_tow = self.vessel_info['Tow'] == 1
        # n_tow_vessels = np.count_nonzero(mask_tow)

        return mask_sp, mask_pile, mask_dea, mask_ml_taut, mask_ml_chain, mask_ten, mask_hu, summation

    def construct_schedule_optimization(self, anchor_id: Union[list, np.ndarray],
                                        ml_type: Union[list, np.ndarray],
                                        vessels: Union[list, np.ndarray],
                                        lengths: Union[list, np.ndarray],
                                        diameters: Union[list, np.ndarray],
                                        work_windows: dict, workability_limits: dict,
                                        proof_load: Union[list, np.ndarray],
                                        shared: Union[list, np.ndarray],
                                        export: bool = False,
                                        name: list = None):
        """

        :return:
        """
        ################################################################################################################
        # ASSERT EVERYTHING GOES WELL
        check = self._check_pop_size()
        mask_sp, mask_pile, mask_dea, mask_ml_taut, mask_ml_chain, mask_ten, mask_hu, summation = check
        assert len(vessels[0]) == summation, 'Check if all vessels are specified in the Excel.'

        if name is not None:
            assert len(name) == len(vessels), 'Name list should be equal to n_pop'

        inter_mob_variant_anchors = np.array([0] * len(vessels))

        ################################################################################################################
        # INITIALIZE SOME VARIABLES
        vessel_times = np.zeros_like(vessels, dtype=float)
        start_time_project = list()
        schedule_input = list()

        extra_ships_times = list()

        n_inter_mobs_anchor_total = list()
        n_inter_mobs_ml_total = list()

        sustainability_data = list()

        ################################################################################################################
        # GET SOME CONSTANTS
        anchors_to_install = list()
        for item in shared:
            if item:  # shared
                anchors_to_install.append(self.project_info['Anchors to install - shared'])
            else:  # not shared
                anchors_to_install.append(self.project_info['Anchors to install'])
        start_year = self.project_info['Start']

        ################################################################################################################
        # DETERMINE ANCHOR TYPE

        anchor_types = ['dea', 'suction', 'suction-integrated', 'pilled-integrated']
        bollard_pull_vessels = self.vessel_info['Bollard pull [kN]'][mask_ml_chain].values

        anchor_to_process = list()
        for it, a_id in enumerate(anchor_id):
            requested_anchor = anchor_types[int(a_id)]
            if requested_anchor == 'dea':
                if min(bollard_pull_vessels) < proof_load[it]:
                    anchor_to_process.append('dea_stev')
                else:
                    anchor_to_process.append('dea_non-stev')
            else:
                anchor_to_process.append(requested_anchor)

        ################################################################################################################
        # GET ALL VESSELS PER TYPE
        s_sp_vessels = 0
        e_sp_vessels = s_sp_vessels + np.count_nonzero(mask_sp)
        sp_install_vessels = vessels[:, s_sp_vessels:e_sp_vessels]
        deck_space_sp = self.vessel_info['Deck length'][mask_sp].values
        names_sp_vessels = self.vessel_info['Vessel'][mask_sp].values

        s_pile_vessels = e_sp_vessels
        e_pile_vessels = s_pile_vessels + np.count_nonzero(mask_pile)
        pile_install_vessels = vessels[:, s_pile_vessels:e_pile_vessels]
        deck_space_piles = self.vessel_info['Deck length'][mask_pile].values
        names_pile_vessels = self.vessel_info['Vessel'][mask_pile].values

        s_dea_vessels = e_pile_vessels
        e_dea_vessels = s_dea_vessels + np.count_nonzero(mask_dea)
        dea_install_vessels = vessels[:, s_dea_vessels:e_dea_vessels]
        deck_space_dea = self.vessel_info['Deck length'][mask_dea].values
        names_dea_vessels = self.vessel_info['Vessel'][mask_dea].values

        s_ml_taut_vessels = e_dea_vessels
        e_ml_taut_vessels = s_ml_taut_vessels + np.count_nonzero(mask_ml_taut)
        ml_taut_install_vessels = vessels[:, s_ml_taut_vessels:e_ml_taut_vessels]
        max_ml_poly_onboard = self.vessel_info['Capacity reels'][mask_ml_taut].values
        names_ml_taut_vessels = self.vessel_info['Vessel'][mask_ml_taut].values

        s_ml_catenary_vessels = e_ml_taut_vessels
        e_ml_catenary_vessels = s_ml_catenary_vessels + np.count_nonzero(mask_ml_chain)
        ml_catenary_install_vessels = vessels[:, s_ml_catenary_vessels:e_ml_catenary_vessels]
        max_ml_chain_onboard = self.vessel_info['Capacity ML'][mask_ml_chain].values
        names_ml_catenary_vessels = self.vessel_info['Vessel'][mask_ml_chain].values

        s_ten_vessels = e_ml_catenary_vessels
        e_ten_vessels = s_ten_vessels + np.count_nonzero(mask_ten)
        tensioning_vessels = vessels[:, s_ten_vessels:e_ten_vessels]
        names_ten_vessels = self.vessel_info['Vessel'][mask_ten].values

        s_hu_vessels = e_ten_vessels
        e_hu_vessels = s_hu_vessels + 1
        hu_vessels = vessels[:, s_hu_vessels:e_hu_vessels]
        names_hu_vessels = self.vessel_info['Vessel'][mask_hu].values

        ################################################################################################################
        # GET WORKABLE MONTHS PER LIMIT AND WORK WINDOW
        work_windows_list = np.unique([work_windows[key] for key in work_windows.keys()])
        workability_dict = workability_limits.copy()
        workability_dict['wind'] = np.unique([workability_dict['wind'][key] for key in workability_dict['wind'].keys()])
        workability_dict['waves'] = np.unique(
            [workability_dict['waves'][key] for key in workability_dict['waves'].keys()])

        # get dict with workable months per workability limit and work window
        percentage_workable_time = 75
        workable_months_per_limit_wind, workable_months_per_limit_waves = self.workability_analysis(
            percentage_workable_time,
            workability_dict,
            work_windows_list)

        ################################################################################################################
        # ANCHOR INSTALL
        # for the iterations, every start year must be set per member of the population
        start_year_pop = np.array([start_year] * len(vessels))

        # set work window
        work_window_anchor_install = work_windows['anchor install']

        # loop through vessel combinations
        start_a_season = list()
        a_installed_per_vessel = list()
        for i in range(len(vessels)):
            self.schedule.append([])
            schedule_input.append([])
            sustainability_data.append(self.s_data_basis.copy())

            extra_ships_times.append(dict())
            ship_times_all = np.zeros(len(vessel_times[0]))

            # set workability limits
            if 'dea' in anchor_to_process[i]:
                workability_limit_id = int(
                    np.where(np.array(workability_dict['waves']) == workability_limits['waves']['anchor install DEA'])[
                        0][0])
            else:
                workability_limit_id = int(
                    np.where(np.array(workability_dict['waves']) == workability_limits['waves']['anchor install'])[0][
                        0])

            # get workable months and set start and end of season
            workable_months_anchor = workable_months_per_limit_waves[workability_limit_id][work_window_anchor_install]

            start_season_anchor_install = datetime.strptime(
                '01/' + month_abbr[workable_months_anchor[0]] + '/' + str(start_year), "%d/%b/%Y")
            start_a_season.append(start_season_anchor_install)

            try:
                end_season_anchor_install = datetime.strptime(
                    '01/' + month_abbr[workable_months_anchor[-1] + 1] + '/' + str(start_year), "%d/%b/%Y")
            except IndexError:  # if workable_months_anchor[-1] == 12
                end_season_anchor_install = datetime.strptime(
                    '31/' + month_abbr[12] + '/' + str(start_year), "%d/%b/%Y")

            if 'suction' in anchor_to_process[i]:
                anchor_install_vessels = sp_install_vessels[i]
                deck_length = deck_space_sp
                s_anchor_vessels = s_sp_vessels
                e_anchor_vessels = e_sp_vessels
                names_anchor_vessels = names_sp_vessels
            elif 'pilled' in anchor_to_process[i]:
                anchor_install_vessels = pile_install_vessels[i]
                deck_length = deck_space_piles
                s_anchor_vessels = s_pile_vessels
                e_anchor_vessels = e_pile_vessels
                names_anchor_vessels = names_pile_vessels
            else:
                anchor_install_vessels = dea_install_vessels[i]
                deck_length = deck_space_dea
                s_anchor_vessels = s_dea_vessels
                e_anchor_vessels = e_dea_vessels
                names_anchor_vessels = names_dea_vessels

            # print()
            # try:
            #     print(f'Anchor {name[i]}')
            # except TypeError:
            #     print('Anchor')
            # print()
            # print(anchor_to_process[i])
            # print(names_anchor_vessels)
            # print('start season:', start_season_anchor_install)
            # print('end season:', end_season_anchor_install)
            result_planner_a = self._anchor_install(anchor_to_process[i], ml_type[i],
                                                    inter_mob_variant_anchors[i],
                                                    anchor_install_vessels,
                                                    deck_length, lengths[i], diameters[i],
                                                    anchors_to_install[i],
                                                    start_season_anchor_install, end_season_anchor_install,
                                                    max_ml_chain_onboard)

            install_duration, ship_times, extra_seasons, mob, demob, n_inter_mobs, t_inter_mobs, s_data, a_per_vessel = result_planner_a
            n_inter_mobs_anchor_total.append([n_inter_mobs, t_inter_mobs])
            a_installed_per_vessel.append(a_per_vessel)

            for key in sustainability_data[i].keys():
                temp_array = np.zeros(len(vessel_times[0]))
                temp_array[s_anchor_vessels:e_anchor_vessels] = s_data[key]
                sustainability_data[i][key] += temp_array

            # set ship times on the locations of the anchor installation vessels
            ship_times_all[s_anchor_vessels:e_anchor_vessels] = ship_times

            # set new vessel times and the start year of the next task
            vessel_times[i] = np.sum([vessel_times[i], ship_times_all], axis=0)

            if '-integrated' not in anchor_to_process[i]:
                start_year_pop[i] += 1 + max(extra_seasons)

            ############################################################################################################
            # Add results to schedule for Gant chart export

            # start is always start season minus mobilization time
            s = start_season_anchor_install - timedelta(days=mob)
            start_time_project.append(s)

            season_length = (end_season_anchor_install - start_season_anchor_install).days
            schedule_input[i].append(
                [ship_times, extra_seasons, season_length, s, mob, demob, names_anchor_vessels, 'Anchor installation'])

        # print()
        ################################################################################################################
        # ML INSTALL

        # set work window
        work_window_ml_install = work_windows['ml install']
        workability_limit_id = int(
            np.where(np.array(workability_dict['waves']) == workability_limits['waves']['ml install'])[0][0])

        # get workable months. Start month and end month are defined in the for loop below
        workable_months_ml = workable_months_per_limit_waves[workability_limit_id][work_window_ml_install]

        # loop through vessel combinations
        for i in range(len(vessels)):
            if anchor_to_process[i] == 'dea_non-stev':
                continue  # ML is part of anchor install and bollard pull > proof load

            # set start and end month based on extra seasons needed for anchor install
            start_season_ml_install = datetime.strptime(
                '01/' + month_abbr[workable_months_ml[0]] + '/' + str(start_year_pop[i]), "%d/%b/%Y")

            if start_season_ml_install < start_a_season[i]:
                start_season_ml_install = start_a_season[i]

            try:
                end_season_ml_install = datetime.strptime(
                    '01/' + month_abbr[workable_months_ml[-1] + 1] + '/' + str(start_year_pop[i]), "%d/%b/%Y")
            except IndexError:  # if workable_months_ml[-1] == 12
                end_season_ml_install = datetime.strptime(
                    '31/' + month_abbr[12] + '/' + str(start_year_pop[i]), "%d/%b/%Y")

            ship_times_all = np.zeros(len(vessel_times[0]))  # set empty array for ship times

            # print()
            # try:
            #     print(f'ML {name[i]}')
            # except TypeError:
            #     print('ML')
            # print()
            # print(ml_type[i])
            # print(names_ml_taut_vessels, names_ml_catenary_vessels, names_ten_vessels)
            # print('start season:', start_season_ml_install)
            # print('end season:', end_season_ml_install)

            if anchor_to_process[i] == 'dea_stev':
                install_duration, ship_times, extra_seasons, mob, demob, s_data = self._tensioning(
                    tensioning_vessels[i],
                    self.project_info['Anchors to install'],
                    start_season_ml_install,
                    end_season_ml_install)

                extra_seasons = [extra_seasons] * len(tensioning_vessels[i])
                n_inter_mobs = [0] * len(ship_times)
                t_inter_mobs = [0] * len(ship_times)

                s_mt_vessels = s_ten_vessels
                e_mt_vessels = e_ten_vessels
                names_mt_vessels = names_ten_vessels

                if 'Scout' in names_mt_vessels:
                    mask = names_mt_vessels == 'Scout'
                    s_data_aht = dict()
                    for it, key in enumerate(s_data.keys()):
                        s_data_aht[key] = s_data[key][mask].copy()
                    extra_ships_times[i]['AHT'] = [ship_times[mask], s_data_aht]
                else:
                    extra_ships_times[i]['HLV'] = [len(ship_times) * [0], self.s_data_basis.copy()]
            else:
                if ml_type[i] == 'taut':
                    ml_install_vessels = ml_taut_install_vessels[i]
                    max_ml_onboard = max_ml_poly_onboard
                else:
                    ml_install_vessels = ml_catenary_install_vessels[i]
                    max_ml_onboard = max_ml_chain_onboard
                result_planner = self._mooring_line_install(anchor_to_process[i], ml_type[i],
                                                            ml_install_vessels,
                                                            max_ml_onboard, self.project_info['Anchors to install'],
                                                            start_season_ml_install, end_season_ml_install,
                                                            a_installed_per_vessel[i])

                install_duration, ship_times, extra_seasons, mob, demob, n_inter_mobs, t_inter_mobs, s_data = result_planner
                extra_ships_times[i]['HLV'] = [len(ship_times) * [0], self.s_data_basis.copy()]

                if ml_type[i] == 'taut':
                    s_mt_vessels = s_ml_taut_vessels
                    e_mt_vessels = e_ml_taut_vessels
                    names_mt_vessels = names_ml_taut_vessels
                else:
                    s_mt_vessels = s_ml_catenary_vessels
                    e_mt_vessels = e_ml_catenary_vessels
                    names_mt_vessels = names_ml_catenary_vessels

            n_inter_mobs_ml_total.append([n_inter_mobs, t_inter_mobs])
            for key in sustainability_data[i].keys():
                temp_array = np.zeros(len(vessel_times[0]))
                temp_array[s_mt_vessels:e_mt_vessels] = s_data[key]
                sustainability_data[i][key] += temp_array

            # set ship times on the locations of the anchor installation vessels
            ship_times_all[s_mt_vessels:e_mt_vessels] = ship_times

            # set new vessel times and the start year of the next task
            vessel_times[i] = np.sum([vessel_times[i], ship_times_all], axis=0)

            ############################################################################################################
            # Add results to schedule for Gant chart export

            # start is always start season minus mobilization time
            s = start_season_ml_install - timedelta(
                days=mob)
            season_length = (end_season_ml_install - start_season_anchor_install).days

            schedule_input[i].append(
                [ship_times, extra_seasons, season_length, s, mob, demob, names_mt_vessels,
                 'Mooring line installation'])

        ############################################################################################################
        # COMPENSATE FOR DIFF NUMBER OF INTER MOBS
        for i in range(len(vessels)):
            if '-integrated' in anchor_to_process[i]:
                if 'suction' in anchor_to_process[i]:
                    s_anchor_vessels = s_sp_vessels
                    e_anchor_vessels = e_sp_vessels
                    a_vessels = sp_install_vessels
                elif 'pilled' in anchor_to_process[i]:
                    s_anchor_vessels = s_pile_vessels
                    e_anchor_vessels = e_pile_vessels
                    a_vessels = pile_install_vessels
                else:
                    s_anchor_vessels = s_dea_vessels
                    e_anchor_vessels = e_dea_vessels
                    a_vessels = dea_install_vessels

                if anchor_to_process[i] == 'dea_stev':
                    s_mt_vessels = s_ten_vessels
                    e_mt_vessels = e_ten_vessels
                    ml_vessels = tensioning_vessels
                else:
                    if ml_type[i] == 'taut':
                        s_mt_vessels = s_ml_taut_vessels
                        e_mt_vessels = e_ml_taut_vessels
                        ml_vessels = ml_taut_install_vessels
                    else:
                        s_mt_vessels = s_ml_catenary_vessels
                        e_mt_vessels = e_ml_catenary_vessels
                        ml_vessels = ml_catenary_install_vessels

                t_inter_mob_anchor = np.multiply(n_inter_mobs_anchor_total[i][0], n_inter_mobs_anchor_total[i][1])
                t_inter_mob_ml = np.multiply(n_inter_mobs_ml_total[i][0], n_inter_mobs_ml_total[i][1])

                mask_a = a_vessels[i] > 0
                mask_ml = ml_vessels[i] > 0
                difference = np.sum(t_inter_mob_anchor) - np.sum(t_inter_mob_ml)
                if difference < 0 and shared[i] is False:
                    schedule_input[i][0][0][mask_a] += -1 * difference / len(np.nonzero(t_inter_mob_ml))
                    temp_array = np.zeros(len(vessel_times[i]))
                    temp_array[s_anchor_vessels:e_anchor_vessels][mask_a] = -1 * difference / len(
                        np.nonzero(t_inter_mob_ml))
                    sustainability_data[i]['idle'] += temp_array
                    vessel_times[i] = np.sum([vessel_times[i], temp_array], axis=0)
                elif difference < 0 and shared[i]:  # n_a << n_ml, use approximation for addition wait time a_vessels
                    approx_difference = self.project_info['Anchors to install - shared'] / self.project_info[
                        'Anchors to install'] * difference
                    schedule_input[i][0][0][mask_a] += -1 * approx_difference / len(np.nonzero(t_inter_mob_ml))
                    temp_array = np.zeros(len(vessel_times[i]))
                    temp_array[s_anchor_vessels:e_anchor_vessels][mask_a] = -1 * approx_difference / len(
                        np.nonzero(t_inter_mob_ml))
                    sustainability_data[i]['idle'] += temp_array
                    vessel_times[i] = np.sum([vessel_times[i], temp_array], axis=0)
                else:
                    schedule_input[i][1][0][mask_ml] += difference / len(np.nonzero(t_inter_mob_anchor))
                    temp_array = np.zeros(len(vessel_times[i]))
                    temp_array[s_mt_vessels:e_mt_vessels][mask_ml] = difference / len(np.nonzero(t_inter_mob_ml))
                    sustainability_data[i]['idle'] += temp_array
                    vessel_times[i] = np.sum([vessel_times[i], temp_array], axis=0)

        ############################################################################################################
        # HOOK-UP AND TOWING

        # set work window
        work_window_hookup = work_windows['hookup']
        work_window_towing = work_windows['towing']

        workability_limit_id_h_wind = int(
            np.where(np.array(workability_dict['wind']) == workability_limits['wind']['hookup'])[0][0])
        workability_limit_id_h_waves = int(
            np.where(np.array(workability_dict['waves']) == workability_limits['waves']['hookup'])[0][0])
        workability_limit_id_t_wind = int(
            np.where(np.array(workability_dict['wind']) == workability_limits['wind']['towing'])[0][0])
        workability_limit_id_t_waves = int(
            np.where(np.array(workability_dict['waves']) == workability_limits['waves']['towing'])[0][0])

        # get workable months per vessel and env. condition
        workable_months_collection = list()
        workable_months_collection.append(workable_months_per_limit_wind[workability_limit_id_h_wind][
                                              work_window_hookup])
        workable_months_collection.append(workable_months_per_limit_waves[workability_limit_id_h_waves][
                                              work_window_hookup])
        workable_months_collection.append(workable_months_per_limit_wind[workability_limit_id_t_wind][
                                              work_window_towing])
        workable_months_collection.append(workable_months_per_limit_waves[workability_limit_id_t_waves][
                                              work_window_towing])

        sizes = [len(item) for item in workable_months_collection]  # list with lengths of workable months lists

        # set turbines to install and check if it is three times smaller than anchor_to_install
        turbines_to_install = self.project_info['Turbines to install']

        # loop through vessel combinations
        for i in range(len(vessels)):
            # set start and end month based on extra seasons needed for anchor install
            workable_months_hookup_towing = workable_months_collection[np.where(np.array(sizes) == min(sizes))[0][0]]

            start_season_hu = datetime.strptime(
                '01/' + month_abbr[workable_months_hookup_towing[0]] + '/' + str(start_year_pop[i]), "%d/%b/%Y")
            try:
                end_season_hu = datetime.strptime(
                    '01/' + month_abbr[workable_months_hookup_towing[-1] + 1] + '/' + str(start_year_pop[i]),
                    "%d/%b/%Y")
            except IndexError:  # if workable_months_anchor[-1] == 12
                end_season_hu = datetime.strptime(
                    '31/' + month_abbr[12] + '/' + str(start_year_pop[i]), "%d/%b/%Y")

            # print()
            # try:
            #     print(f'HU {name[i]}')
            # except TypeError:
            #     print('HU')
            # print()
            # print(names_hu_vessels)
            # print('start season:', start_season_hu)
            # print('end season:', end_season_hu)

            ship_times_all = np.zeros(len(vessel_times[0]))  # set empty array for ship times
            install_duration, ship_times, ship_times_tow, extra_seasons, mob, demob, s_data, s_data_tow = self._hookup_towing(
                hu_vessels[i],
                turbines_to_install,
                start_season_hu,
                end_season_hu)

            extra_ships_times[i]['tugs'] = [ship_times_tow, s_data_tow]
            extra_seasons = [extra_seasons] * len(hu_vessels[i])

            # set ship times on the locations of the anchor installation vessels
            ship_times_all[s_hu_vessels:e_hu_vessels] = ship_times

            for key in sustainability_data[i].keys():
                temp_array = np.zeros(len(vessel_times[0]))
                temp_array[s_hu_vessels:e_hu_vessels] = s_data[key]
                sustainability_data[i][key] += temp_array

            # set new vessel times and the start year of the next task
            vessel_times[i] = np.sum([vessel_times[i], ship_times_all], axis=0)
            try:
                start_year_pop[i] += 1 + max(extra_seasons)
            except OverflowError:
                start_year_pop[i] += 1 + 5

            ############################################################################################################
            # Add results to schedule for Gant chart export

            # start is always start season minus mobilization time
            s = start_season_hu - timedelta(
                days=sum(self.hookup_schedule.building_blocks['mobilization HUV']['time']))
            season_length = (end_season_hu - start_season_hu).days

            ship_times_scheduler = [0] * len(mask_hu)
            extra_seasons_export = [0] * len(mask_hu)
            ship_times_scheduler[int(hu_vessels[i])] = ship_times[0]
            extra_seasons_export[int(hu_vessels[i])] = extra_seasons[0]
            schedule_input[i].append(
                [ship_times_scheduler, extra_seasons_export, season_length, s, mob, demob, names_hu_vessels, 'Hookup'])

        end_time_project = self._append_schedule(schedule_input)

        project_duration = list()
        for i in range(len(start_time_project)):
            project_duration.append((end_time_project[i] - start_time_project[i]).days)
        if export:
            for i in range(len(hu_vessels)):
                fig = ff.create_gantt(self.schedule[i][::-1], index_col='Resource', show_colorbar=True,
                                      showgrid_x=True, showgrid_y=True, group_tasks=True)
                if name is not None:
                    fig.write_html(f'./schedule/html_exports/{i + 1}_schedule_{name[i]}_{anchor_to_process[i]}.html')
                else:
                    fig.write_html(f'./schedule/html_exports/{i + 1}_schedule_{anchor_to_process[i]}.html')

        return project_duration, vessel_times, extra_ships_times, sustainability_data

    def _append_schedule(self, inp):
        end_times = list()

        for i in range(len(inp)):
            for item in inp[i]:
                ship_times, extra_seasons, season_length, start_season, mob, demob, names_vessels, task = item
                # loop through the times a ship is needed
                for it, time in enumerate(ship_times):
                    if time == 0.:
                        continue

                    if season_length >= 364:
                        extra_seasons[it] = 0

                    if extra_seasons[it] < 0:
                        extra_seasons[it] = 0

                    pure_vessel_time = time - (extra_seasons[it] + 1) * (mob + demob)
                    pure_vessel_time_per_season = pure_vessel_time / (extra_seasons[it] + 1)
                    s = start_season + timedelta(mob)
                    if extra_seasons[it] > 0:  # if more seasons than 1 are needed
                        # consider the time still to go
                        days_to_go = pure_vessel_time
                        while 1:
                            days_to_go_old = days_to_go
                            days_to_go -= season_length
                            if days_to_go < 0:
                                days_to_go = days_to_go_old
                                break
                            s_mob = s - timedelta(days=mob)
                            e_mob = s
                            self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                         Start=s_mob,
                                                         Finish=e_mob,
                                                         Resource='Mobilization'))
                            s_task = s
                            e_task = s_task + timedelta(days=pure_vessel_time_per_season)
                            self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                         Start=s_task,
                                                         Finish=e_task,
                                                         Resource=task))
                            s_demob = e_task
                            e_demob = s_demob + timedelta(days=demob)
                            self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                         Start=s_demob,
                                                         Finish=e_demob,
                                                         Resource='Demobilization'))
                            s = s + timedelta(days=365)

                        # add duration in final season
                        s_mob = s - timedelta(days=mob)
                        e_mob = s
                        self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                     Start=s_mob,
                                                     Finish=e_mob,
                                                     Resource='Mobilization'))
                        s_task = s
                        e_task = s_task + timedelta(days=days_to_go)
                        self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                     Start=s_task,
                                                     Finish=e_task,
                                                     Resource=task))
                        s_demob = e_task
                        e_demob = s_demob + timedelta(days=demob)
                        self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                     Start=s_demob,
                                                     Finish=e_demob,
                                                     Resource='Demobilization'))

                    else:  # if one season is enough
                        s_mob = s - timedelta(days=mob)
                        e_mob = s
                        self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                     Start=s_mob,
                                                     Finish=e_mob,
                                                     Resource='Mobilization'))
                        s_task = s
                        try:
                            e_task = s_task + timedelta(days=pure_vessel_time)
                        except ValueError as err:
                            e_task = s_task + timedelta(days=1e4)
                        self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                     Start=s_task,
                                                     Finish=e_task,
                                                     Resource=task))
                        s_demob = e_task
                        e_demob = s_demob + timedelta(days=demob)
                        self.schedule[i].append(dict(Task=f'{names_vessels[it]}',
                                                     Start=s_demob,
                                                     Finish=e_demob,
                                                     Resource='Demobilization'))
                if task == 'Hookup':
                    end_times.append(e_demob)
        return end_times


if __name__ == '__main__':
    cls = CreateSchedule()
    vd = cls._extract_data('Vessels')['vessel_data'].values

    wl = {
        'wind': {
            'anchor install': 15,
            'anchor install DEA': 20,
            'ml install': 15,
            'towing': 20,
            'hookup': 12
        },
        'waves': {
            'anchor install': 2,
            'anchor install DEA': 2.5,
            'ml install': 2.5,
            'towing': 3.5,
            'hookup': 2
        }
    }

    ww = {  # hours
        'anchor install': 12,  # suction anchor install
        'ml install': 12,
        'towing': 54,
        'hookup': 54,
    }

    vessel = np.array([
        [0., 1., 0., 0., 0., 0., 1., 0., 1., 0., 0., 0., 0., 1., 0., 0., 1., 1., 1., 1., 0., 0., 3.],
        [0., 1., 0., 0., 0., 0., 0., 0., 0., 1., 1., 1., 0., 1., 0., 0., 0., 0., 0., 1., 0., 0., 0.]
        # [1, 0, 0, 0, 1, 0, 1, 0, 0, 1, 0, 0, 0, 0, 0, 0, 1, 0, 1, 0, 0, 0, 2],
    ])

    inter_m = np.array([0, 0])
    at = ['dea', 'suction', 'suction-integrated', 'pilled-integrated']
    a = [2, 1]
    ml_t = ['taut', 'taut']

    pd, vt, est, sd, nt = cls.construct_schedule_optimization(a, ml_t, vessel, [29, 39], [4.62, 5.1], ww, wl,
                                                              proof_load=[4800, 4800],
                                                              export=True, shared=[1, 1])

    pprint(vt)
    # for i in range(len(vt)):
    #     pprint(est[i])
    #     print(f'Ship times {i + 1}: {np.sum(vt[i])}')
    #     a = 0
    #     for key in sd[i].keys():
    #         a += np.sum(sd[i][key])
    #     print(f'SD times {i + 1}: {a}')
